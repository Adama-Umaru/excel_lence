import openpyxl

# Read data from an Excel file
workbook = openpyxl.load_workbook("data.xlsx")
sheet = workbook.active
for row in sheet.iter_rows(values_only=True):
    print(row)

# # Write data to an Excel file
workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.title = "Data"
sheet["A1"] = "Name"
sheet["B1"] = "Age"
sheet["C1"] = "Gender"
sheet.append(("Alice", 25, "Female"))
sheet.append(("Bob", 30, "Male"))
workbook.save("output.xlsx")

# # Convert a text file to an Excel file
with open("data.txt", "r") as f:
    data = f.read().splitlines()
workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.title = "Data"
for i, row in enumerate(data):
    sheet.cell(row=i+1, column=1, value=row)
workbook.save("data.xlsx")




